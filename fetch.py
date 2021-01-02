# -*- coding: utf8 -*-
import requests
import time
from datetime import datetime, timedelta
import lxml.html
import openpyxl


ALADIN = 'https://www.aladin.co.kr'
ALADIN_LOGIN = ALADIN + '/login/wlogin_popup.aspx'
BOOKPLE = 'http://blog.aladin.co.kr'
BOOKPLE_LIST = BOOKPLE + '/ucl/bookple/ajax/listRepeator_ajax.aspx'
BOOKPLE_XPATHS = {
    'FEED': '//div[@class="feed_one2"]',
    'COVER': './div[@class="feed_recm_coverbox1"]/img',
    'TITLE': './/div[@class="viewpage_coment"]//li[1]/span',
    'AUTHORS': './/div[@class="viewpage_coment"]//li[2]',
    'MAX_PAGE': '//input[@id="MaxPageCount"]',
}
BOOKPLE_TYPES = {
    'WISH': dict(type='ItemWish', status= 4),
    'READING': dict(type='ItemReading', status=3),
    'READED': dict(type='ItemReadCompleted', status=1),
}

def login_aladin(config):
    session = requests.Session()
    data = dict(Email=config['bookple']['id'],
                Password=config['bookple']['password'],
                Action=1,
                ReturnUrl=None,
                ReturnUrl_pop=None,
                SecureLogin=False,
                snsUserId=0,
                snsType=0,
                snsAppId=1)
    res = session.post(ALADIN_LOGIN, data=data)
    return session

def pager(session, page=0, types=None):
    if not types:
        types = BOOKPLE_TYPES['WISH']
    timestamp = int(round(time.time() * 1000))
    params = dict(tstamp=int(round(time.time() * 1000)),
                  page=page,
                  BookplePaperApi=types['type'],
                  IsContentsView=1,
                  CurrentBlogID=742408194,
                  ViewRowCount=10,
                  ReadStatus=types['status'],
                  AuthorId=0,
                  SeriesId=0,
                  CID=0,
                  ItemId=0)
    res = session.get(BOOKPLE_LIST, params=params)
    elem = lxml.html.fromstring(res.text)
    maxpage = int(elem.xpath(BOOKPLE_XPATHS['MAX_PAGE'])[0].get('value', '0'))
    if maxpage < page:
        raise ValueError('exceed max page')
    # TODO: max page count
    return elem

def items(page):
    xpath = page.xpath
    feeds = xpath(BOOKPLE_XPATHS['FEED'])
    for feed in feeds:
        img = feed.xpath(BOOKPLE_XPATHS['COVER'])[0].get('src')
        title = (feed.xpath(BOOKPLE_XPATHS['TITLE'])[0].text or '')
        authors = (feed.xpath(BOOKPLE_XPATHS['AUTHORS'])[0].text or '')
        yield dict(title=title, authors=authors, image=img)

def find_from_worksheet(item, ws):
    _skip_first_line = False
    for row in ws.iter_rows():
        if not _skip_first_line:
            _skip_first_line = True
            continue
        if (row[0].value or '') == item['title'] and \
                (row[1].value or '') == item['authors']:
            return row[0].row

def item_to_worksheet(item, ws, prevs=None):
    if not prevs:
        prevs = []
    row = find_from_worksheet(item, ws)
    if row:
        return
    for p in prevs:
        row = find_from_worksheet(item, p)
        if not row:
            continue
        #print 'remove row %d from %s: %s' % (row, p.title, item['title'])
        p.remove_rows.append(row)
    ws.append((item['title'], item['authors'], item['image']))

if __name__ == '__main__':
    import yaml

    config = yaml.load(open('config.yml'))
    aladin_session = login_aladin(config)
    try:
        workbook = openpyxl.load_workbook(config['xlsx']['file'])
    except IOError:
        workbook = openpyxl.Workbook()
    prevs = []
    maxpage = config['page']
    if type(maxpage) == str and maxpage.lower() == 'full':
        maxpage = 0
    for title, types in BOOKPLE_TYPES.items():
        try:
            ws = workbook.get_sheet_by_name(title)
        except KeyError:
            ws = workbook.create_sheet(title=title)
            ws.cell(row=1, column=1, value='TITLE')
            ws.cell(row=1, column=2, value='AUTHORS')
            ws.cell(row=1, column=3, value='IMAGE_URL')
        ws.remove_rows = []
        page = 1
        while True:
            try:
                for item in items(pager(aladin_session, page, types=types)):
                    item_to_worksheet(item, ws, prevs=prevs)
            except ValueError:
                break
            page += 1
            if maxpage <= 0:
                continue
            if maxpage < page:
                break
        prevs.append(ws)
    for ws in prevs:
        workbook.remove_sheet(ws)
        new_sheet = workbook.create_sheet(title=ws.title)
        new_sheet.column_dimensions['A'].width = 50
        new_sheet.column_dimensions['B'].width = 50
        new_sheet.column_dimensions['C'].width = 50
        for idx, row in enumerate(ws.iter_rows()):
            if idx + 1 in ws.remove_rows:
                continue
            if row[0].value == None:
                continue
            new_sheet.append([c.value for c in row])
        new_sheet.title = ws.title
    workbook.save(config['xlsx']['file'])
